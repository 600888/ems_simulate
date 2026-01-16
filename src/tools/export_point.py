import csv

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class PointExporter:
    def __init__(self, device, file_path: str) -> None:
        self.device = device
        self.file_path = file_path

    # 导出xlsx文件
    def exportDataPointXlsx(self, file_path: str) -> None:
        # 创建一个工作簿
        wb = Workbook()

        # 获取默认的工作表
        default_ws = wb.active

        # 删除默认的工作表
        wb.remove(default_ws)

        head = [
            "从机地址",
            "地址",
            "16进制地址",
            "功能码",
            "测点名称",
            "测点编码",
            "测点值",
            "16进制测点值",
            "乘积系数",
            "加法系数",
            "帧类型",
        ]

        for i in range(0, self.device.modbus_tcp_server.slave_cnt):
            slave_id = i + 1
            yc_list = self.device.yc_dict.get(slave_id)
            yx_list = self.device.yx_dict.get(slave_id)
            frame_type_dict = self.device.frame_type_dict()

            # 创建一个工作表
            if slave_id == 1:
                ws = wb.create_sheet("主机" + str(slave_id))
                ws.title = "主机" + str(slave_id)
            else:
                ws = wb.create_sheet("电池簇" + str(slave_id))
                ws.title = "电池簇" + str(slave_id)

            # 创建一个填充对象
            yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
            green = PatternFill(fill_type="solid", fgColor="00FF00")
            red = PatternFill(fill_type="solid", fgColor="FF0000")
            blue = PatternFill(fill_type="solid", fgColor="0000FF")

            row = 1  # 记录行数
            cnt = 0  # 用来判断是否为第一行
            if len(yc_list) > 0:
                for yc in yc_list:
                    if cnt == 0:
                        ws.append(["遥测"])
                        # 合并单元格
                        ws.merge_cells("A%d:K%d" % (row, row))
                        # 设置单元格的颜色
                        ws["A%d" % row].fill = blue
                        row += 1
                        # 设置表头
                        ws.append(head)
                    row += 1
                    cnt += 1
                    ws.append(
                        [
                            yc.rtu_addr,
                            yc.address,
                            yc.hex_address,
                            yc.func_code,
                            yc.name,
                            yc.code,
                            yc.value,
                            yc.hex_value,
                            yc.mul_coe,
                            yc.add_coe,
                            frame_type_dict[yc.frame_type],
                        ]
                    )
                    ws["G%d" % row].fill = yellow

                # 空一行
                ws.append([])
                cnt = 0
                row += 2

            if len(yx_list) > 0:
                for yx in yx_list:
                    if cnt == 0:
                        ws.append(["遥信"])
                        # 合并单元格
                        ws.merge_cells("A%d:K%d" % (row, row))
                        # 设置单元格的颜色
                        ws["A%d" % row].fill = green
                        row += 1
                        # 设置表头
                        ws.append(head)
                    row += 1
                    cnt += 1
                    ws.append(
                        [
                            yx.rtu_addr,
                            yx.address,
                            yx.hex_address,
                            yx.func_code,
                            yx.name,
                            yx.code,
                            yx.value,
                            yx.hex_value,
                            "",
                            "",
                            frame_type_dict[yx.frame_type],
                        ]
                    )
                    ws["G%d" % row].fill = yellow
                # 空一行
                ws.append([])

                # 创建一个对齐对象
                alignment = Alignment(horizontal="center", vertical="center")
                # 创建一个边框对象
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

            # 设置单元格的对齐方式
            for row in ws.iter_rows():
                for cell in row:
                    if cell is not None:
                        cell.alignment = alignment
                        cell.border = thin_border

            # 计算每列的最大长度
            max_length = {}
            # 将max_length 初始化为表头的宽度
            for i, cell in enumerate(head):
                max_length[i] = len(cell) * 2
            for row in self.device.yc_dict[slave_id] + self.device.yx_dict[slave_id]:
                for i, cell in enumerate(row.list()):
                    if len(str(cell)) > max_length.get(i, 0):
                        max_length[i] = len(str(cell))

            # 设置列的宽度
            for i, length in max_length.items():
                ws.column_dimensions[get_column_letter(i + 1)].width = length

        # 保存文件
        wb.save(file_path)

    def exportDataPointCsv(self, file_path):
        open(file_path, "w").close()
        frame_type_dict = self.device.frame_type_dict()
        # 跟据从机地址导出csv文件
        with open(file_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "从机地址",
                    "地址",
                    "位",
                    "功能码",
                    "测点名称",
                    "测点编码",
                    "测点值",
                    "测点值(16进制)",
                    "乘法系数",
                    "加法系数",
                    "帧类型",
                ]
            )
            for slave_id in range(0, len(self.device.yc_dict)):
                if len(self.device.yc_dict[slave_id]) > 0:
                    for yc in self.device.yc_dict[slave_id]:
                        # 全部导出为str格式
                        frame_type = frame_type_dict[yc.frame_type]
                        writer.writerow(
                            [
                                yc.rtu_addr,
                                yc.address,
                                yc.func_code,
                                yc.name,
                                yc.code,
                                str(yc.value),
                                str(yc.hex_value),
                                str(yc.mul_coe),
                                str(yc.add_coe),
                                frame_type,
                            ]
                        )
                    # 空一行
                    writer.writerow([])

                if len(self.device.yx_dict[slave_id]) > 0:
                    for yx in self.device.yx_dict[slave_id]:
                        frame_type = frame_type_dict[yx.frame_type]
                        writer.writerow(
                            [
                                yx.rtu_addr,
                                yx.address,
                                yx.bit,
                                yx.func_code,
                                yx.name,
                                yx.code,
                                str(yx.value),
                                str(yx.hex_value),
                                "",
                                "",
                                frame_type,
                            ]
                        )
                    # 空一行
                    writer.writerow([])
