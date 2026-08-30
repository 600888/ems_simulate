"""
Excel 点表导入模块
支持从 Excel 文件导入四类测点（遥测/遥信/遥控/遥调）
Excel 需包含 4 个 sheet: 遥测, 遥信, 遥控, 遥调
"""

import json
import os
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.data.controller.db import local_session
from src.data.log import log
from src.data.model.point_yc import PointYc
from src.data.model.point_yk import PointYk
from src.data.model.point_yt import PointYt
from src.data.model.point_yx import PointYx
from src.enums.modbus_register import Decode
from src.proto.dnp3.point_config import Dnp3PointConfig


class ExcelPointImporter:
    """Excel 点表导入器"""

    # Sheet 名称映射
    SHEET_NAMES = {
        "yc": "遥测",
        "yx": "遥信",
        "yk": "遥控",
        "yt": "遥调",
    }

    DNP3_POINT_TYPES = {
        0: "Analog Input",
        1: "Binary Input",
        2: "Binary Output",
        3: "Analog Output",
    }
    DNP3_POINT_TYPE_ALIASES = {
        0: {"analog input", "analoginput", "模拟量输入", "遥测"},
        1: {"binary input", "binaryinput", "开关量输入", "遥信"},
        2: {"binary output", "binaryoutput", "开关量输出", "遥控"},
        3: {"analog output", "analogoutput", "模拟量输出", "遥调"},
    }
    DNP3_CONFIG_HEADERS = {
        "static_variation": ("静态变体", "DNP3静态变体"),
        "event_variation": ("事件变体", "DNP3事件变体"),
        "event_class": ("CLASS", "事件类别", "DNP3事件类别"),
        "deadband": ("死区", "DNP3死区"),
        "control_mode": ("控制模式", "DNP3控制模式"),
        "crob_operation": ("CROB操作", "CROB操作类型"),
        "pulse_on_ms": ("脉冲导通时间(ms)", "脉冲导通时间"),
        "pulse_off_ms": ("脉冲断开时间(ms)", "脉冲断开时间"),
        "pulse_count": ("脉冲次数",),
        "initial_quality": ("初始品质", "DNP3初始品质"),
        "event_enabled": ("产生事件", "事件启用"),
        "timestamp_enabled": ("事件携带时标", "时标启用"),
    }

    # 遥测表头
    YC_HEADERS = [
        "code",  # 测点编码
        "name",  # 测点名称
        "rtu_addr",  # 从机地址
        "reg_addr",  # 寄存器地址
        "func_code",  # 功能码
        "decode_code",  # 解析码
        "mul_coe",  # 乘系数
        "add_coe",  # 加系数
        "max_limit",  # 上限值
        "min_limit",  # 下限值
        "iec104_type",  # 104ASDU类型
    ]

    # 遥信表头
    YX_HEADERS = [
        "code",  # 测点编码
        "name",  # 测点名称
        "rtu_addr",  # 从机地址
        "reg_addr",  # 寄存器地址
        "func_code",  # 功能码
        "decode_code",  # 解析码
        "bit",  # 位偏移
        "reverse",  # 是否反转
        "iec104_type",  # 104ASDU类型
    ]

    # 遥控表头
    YK_HEADERS = [
        "code",  # 测点编码
        "name",  # 测点名称
        "rtu_addr",  # 从机地址
        "reg_addr",  # 寄存器地址
        "func_code",  # 功能码
        "decode_code",  # 解析码
        "bit",  # 位偏移
        "command_type",  # 命令类型
        "related_yx_code",  # 关联遥信编码
        "iec104_type",  # 104ASDU类型
    ]

    # 遥调表头
    YT_HEADERS = [
        "code",  # 测点编码
        "name",  # 测点名称
        "rtu_addr",  # 从机地址
        "reg_addr",  # 寄存器地址
        "func_code",  # 功能码
        "decode_code",  # 解析码
        "mul_coe",  # 乘系数
        "add_coe",  # 加系数
        "max_limit",  # 上限值
        "min_limit",  # 下限值
        "related_yc_code",  # 关联遥测编码
        "iec104_type",  # 104ASDU类型
    ]

    def __init__(self, channel_id: int):
        """
        Args:
            channel_id: 通道ID，导入的测点将关联到此通道
        """
        self.channel_id = channel_id
        self.yc_count = 0
        self.yx_count = 0
        self.yk_count = 0
        self.yt_count = 0

    @staticmethod
    def _header_indices(sheet: Worksheet) -> dict[str, int]:
        """返回去除首尾空白后的表头到列下标映射。"""
        return {
            str(cell.value).strip(): index
            for index, cell in enumerate(sheet[1])
            if cell.value is not None and str(cell.value).strip()
        }

    @classmethod
    def _value_by_header(cls, sheet: Worksheet, row: tuple[Any, ...], *headers: str) -> Any:
        indices = cls._header_indices(sheet)
        for header in headers:
            index = indices.get(header)
            if index is not None and index < len(row):
                return row[index]
        return None

    @classmethod
    def _iec104_type(cls, sheet: Worksheet, row: tuple[Any, ...]) -> str | None:
        value = cls._value_by_header(sheet, row, "104ASDU类型", "IEC104类型")
        return str(value).strip() if value is not None and str(value).strip() else None

    @staticmethod
    def _parse_bool(value: Any, field_label: str) -> bool:
        if type(value) in (int, float) and value in (0, 1):
            return bool(value)
        raise ValueError(f"{field_label}必须填写数值 0 或 1")

    @staticmethod
    def _parse_variation(value: Any) -> int:
        if isinstance(value, (int, float)):
            return int(value)
        normalized = str(value).strip().upper()
        if normalized.startswith("V"):
            normalized = normalized[1:]
        return int(normalized)

    @staticmethod
    def _parse_event_class(value: Any) -> int:
        match = re.fullmatch(r"(?:class\s*)?([1-3])", str(value).strip(), flags=re.IGNORECASE)
        if not match:
            raise ValueError("CLASS必须是 Class 1、Class 2 或 Class 3")
        return int(match.group(1))

    @classmethod
    def _parse_dnp3_config(
        cls,
        sheet: Worksheet,
        row: tuple[Any, ...],
        frame_type: int,
        excel_row: int,
    ) -> str | None:
        """从 DNP3 专用列构造可持久化的点级配置 JSON。"""
        headers = cls._header_indices(sheet)
        dnp3_headers = {"DNP3点位类型"}
        for aliases in cls.DNP3_CONFIG_HEADERS.values():
            dnp3_headers.update(aliases)
        if not dnp3_headers.intersection(headers):
            return None

        point_type = cls._value_by_header(sheet, row, "DNP3点位类型")
        if point_type is not None and str(point_type).strip():
            normalized_type = re.sub(r"\s*\([^)]*\)\s*$", "", str(point_type)).strip().lower()
            if normalized_type not in cls.DNP3_POINT_TYPE_ALIASES[frame_type]:
                expected = cls.DNP3_POINT_TYPES[frame_type]
                raise ValueError(f"{sheet.title}第{excel_row}行DNP3点位类型应为 {expected}")

        values: dict[str, Any] = {}
        for field, aliases in cls.DNP3_CONFIG_HEADERS.items():
            value = cls._value_by_header(sheet, row, *aliases)
            if value is None or str(value).strip() == "":
                continue
            if field in {"static_variation", "event_variation"}:
                values[field] = cls._parse_variation(value)
            elif field == "event_class":
                values[field] = cls._parse_event_class(value)
            elif field == "deadband":
                values[field] = float(value)
            elif field in {"pulse_on_ms", "pulse_off_ms", "pulse_count"}:
                values[field] = int(value)
            elif field == "initial_quality":
                values[field] = int(value) if isinstance(value, (int, float)) else int(str(value), 0)
            elif field in {"event_enabled", "timestamp_enabled"}:
                try:
                    values[field] = cls._parse_bool(value, aliases[0])
                except ValueError as exc:
                    raise ValueError(f"{sheet.title}第{excel_row}行DNP3配置无效: {exc}") from exc
            else:
                values[field] = str(value).strip().lower()

        try:
            config = Dnp3PointConfig.from_mapping(frame_type, values)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{sheet.title}第{excel_row}行DNP3配置无效: {exc}") from exc
        return json.dumps(config.to_dict(), ensure_ascii=False, separators=(",", ":"))

    def _clear_existing_points(self) -> None:
        """清除该通道已有的测点数据"""
        from src.data.controller.db import local_session
        from src.data.model.point_yc import PointYc
        from src.data.model.point_yk import PointYk
        from src.data.model.point_yt import PointYt
        from src.data.model.point_yx import PointYx

        try:
            with local_session() as session, session.begin():
                session.query(PointYc).where(PointYc.channel_id == self.channel_id).delete()
                session.query(PointYx).where(PointYx.channel_id == self.channel_id).delete()
                session.query(PointYk).where(PointYk.channel_id == self.channel_id).delete()
                session.query(PointYt).where(PointYt.channel_id == self.channel_id).delete()
            log.info(f"已清除通道 {self.channel_id} 的旧测点数据")
        except Exception as e:
            log.error(f"清除旧测点数据失败: {e}")
            raise e

    def import_from_excel(self, file_path: str) -> tuple[int, int, int, int]:
        """从 Excel 导入测点

        Args:
            file_path: Excel 文件路径

        Returns:
            (yc_count, yx_count, yk_count, yt_count) 各类型导入数量
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        # 先删除该通道的旧测点数据，避免重复导入时UNIQUE约束冲突
        self._clear_existing_points()

        wb = load_workbook(file_path)

        # 导入四类测点
        if self.SHEET_NAMES["yc"] in wb.sheetnames:
            self._import_yc(wb[self.SHEET_NAMES["yc"]])
        if self.SHEET_NAMES["yx"] in wb.sheetnames:
            self._import_yx(wb[self.SHEET_NAMES["yx"]])
        if self.SHEET_NAMES["yk"] in wb.sheetnames:
            self._import_yk(wb[self.SHEET_NAMES["yk"]])
        if self.SHEET_NAMES["yt"] in wb.sheetnames:
            self._import_yt(wb[self.SHEET_NAMES["yt"]])

        wb.close()

        log.info(
            f"Excel导入完成: 遥测={self.yc_count}, 遥信={self.yx_count}, 遥控={self.yk_count}, 遥调={self.yt_count}"
        )
        return (self.yc_count, self.yx_count, self.yk_count, self.yt_count)

    def _import_yc(self, sheet: Worksheet) -> None:
        """导入遥测点"""
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        with local_session() as session, session.begin():
            for excel_row, row in enumerate(rows, start=2):
                if not row[0]:  # 跳过空行
                    continue
                decode_code = str(row[5]) if row[5] else "0x41"
                mul_coe = float(row[6]) if row[6] else 1.0
                add_coe = float(row[7]) if row[7] else 0.0
                calc_max, calc_min = Decode.get_limits_by_code(decode_code, mul_coe, add_coe)

                point = PointYc(
                    code=str(row[0]),
                    name=str(row[1]) if row[1] else "",
                    channel_id=self.channel_id,
                    rtu_addr=int(row[2]) if row[2] else 1,
                    reg_addr=str(row[3]) if row[3] else "0x0000",
                    func_code=int(row[4]) if row[4] else 3,
                    decode_code=decode_code,
                    mul_coe=mul_coe,
                    add_coe=add_coe,
                    max_limit=float(row[8]) if row[8] is not None and str(row[8]).strip() != "" else calc_max,
                    min_limit=float(row[9])
                    if len(row) > 9 and row[9] is not None and str(row[9]).strip() != ""
                    else calc_min,
                    iec_type_id=self._iec104_type(sheet, row),
                    dnp3_config=self._parse_dnp3_config(sheet, row, 0, excel_row),
                )
                session.add(point)
                self.yc_count += 1

    def _import_yx(self, sheet: Worksheet) -> None:
        """导入遥信点"""
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        with local_session() as session, session.begin():
            for excel_row, row in enumerate(rows, start=2):
                if not row[0]:
                    continue
                point = PointYx(
                    code=str(row[0]),
                    name=str(row[1]) if row[1] else "",
                    channel_id=self.channel_id,
                    rtu_addr=int(row[2]) if row[2] else 1,
                    reg_addr=str(row[3]) if row[3] else "0x0000",
                    func_code=int(row[4]) if row[4] else 1,
                    decode_code=str(row[5]) if row[5] else "0x20",
                    bit=int(row[6]) if row[6] else None,
                    reverse=bool(row[7]) if len(row) > 7 and row[7] else False,
                    iec_type_id=self._iec104_type(sheet, row),
                    dnp3_config=self._parse_dnp3_config(sheet, row, 1, excel_row),
                )
                session.add(point)
                self.yx_count += 1

    def _import_yk(self, sheet: Worksheet) -> None:
        """导入遥控点"""
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        with local_session() as session, session.begin():
            for excel_row, row in enumerate(rows, start=2):
                if not row[0]:
                    continue
                point = PointYk(
                    code=str(row[0]),
                    name=str(row[1]) if row[1] else "",
                    channel_id=self.channel_id,
                    rtu_addr=int(row[2]) if row[2] else 1,
                    reg_addr=str(row[3]) if row[3] else "0x0000",
                    func_code=int(row[4]) if row[4] else 5,
                    decode_code=str(row[5]) if row[5] else "0x20",
                    bit=int(row[6]) if row[6] else None,
                    command_type=int(row[7]) if len(row) > 7 and row[7] else 0,
                    # related_yx_id 需要后续通过 code 查找
                    iec_type_id=self._iec104_type(sheet, row),
                    dnp3_config=self._parse_dnp3_config(sheet, row, 2, excel_row),
                )
                session.add(point)
                self.yk_count += 1

    def _import_yt(self, sheet: Worksheet) -> None:
        """导入遥调点"""
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        with local_session() as session, session.begin():
            for excel_row, row in enumerate(rows, start=2):
                if not row[0]:
                    continue
                decode_code = str(row[5]) if row[5] else "0x41"
                mul_coe = float(row[6]) if row[6] else 1.0
                add_coe = float(row[7]) if row[7] else 0.0
                calc_max, calc_min = Decode.get_limits_by_code(decode_code, mul_coe, add_coe)

                point = PointYt(
                    code=str(row[0]),
                    name=str(row[1]) if row[1] else "",
                    channel_id=self.channel_id,
                    rtu_addr=int(row[2]) if row[2] else 1,
                    reg_addr=str(row[3]) if row[3] else "0x0000",
                    func_code=int(row[4]) if row[4] else 6,
                    decode_code=decode_code,
                    mul_coe=mul_coe,
                    add_coe=add_coe,
                    max_limit=float(row[8])
                    if len(row) > 8 and row[8] is not None and str(row[8]).strip() != ""
                    else calc_max,
                    min_limit=float(row[9])
                    if len(row) > 9 and row[9] is not None and str(row[9]).strip() != ""
                    else calc_min,
                    # related_yc_id 需要后续通过 code 查找
                    iec_type_id=self._iec104_type(sheet, row),
                    dnp3_config=self._parse_dnp3_config(sheet, row, 3, excel_row),
                )
                session.add(point)
                self.yt_count += 1


def create_excel_template(file_path: str) -> None:
    """创建 Excel 模板文件

    Args:
        file_path: 输出文件路径
    """
    wb = Workbook()

    # 遥测 sheet
    ws_yc = wb.active
    ws_yc.title = "遥测"
    ws_yc.append(
        [
            "测点编码",
            "测点名称",
            "从机地址",
            "寄存器地址",
            "功能码",
            "解析码",
            "乘系数",
            "加系数",
            "上限值",
            "下限值",
            "104ASDU类型",
        ]
    )

    # 遥信 sheet
    ws_yx = wb.create_sheet("遥信")
    ws_yx.append(
        ["测点编码", "测点名称", "从机地址", "寄存器地址", "功能码", "解析码", "位偏移", "是否反转", "104ASDU类型"]
    )

    # 遥控 sheet
    ws_yk = wb.create_sheet("遥控")
    ws_yk.append(
        [
            "测点编码",
            "测点名称",
            "从机地址",
            "寄存器地址",
            "功能码",
            "解析码",
            "位偏移",
            "命令类型",
            "关联遥信编码",
            "104ASDU类型",
        ]
    )

    # 遥调 sheet
    ws_yt = wb.create_sheet("遥调")
    ws_yt.append(
        [
            "测点编码",
            "测点名称",
            "从机地址",
            "寄存器地址",
            "功能码",
            "解析码",
            "乘系数",
            "加系数",
            "上限值",
            "下限值",
            "关联遥测编码",
            "104ASDU类型",
        ]
    )

    wb.save(file_path)
    wb.close()
    print(f"模板已创建: {file_path}")


def create_sample_excel(file_path: str, protocol: str = "modbus") -> None:
    """创建示例 Excel 文件

    Args:
        file_path: 输出文件路径
        protocol: 协议类型 (modbus/iec104/dlt645)
    """
    wb = Workbook()

    # ===== 遥测 sheet =====
    ws_yc = wb.active
    ws_yc.title = "遥测"
    ws_yc.append(
        [
            "测点编码",
            "测点名称",
            "从机地址",
            "寄存器地址",
            "功能码",
            "解析码",
            "乘系数",
            "加系数",
            "上限值",
            "下限值",
            "104ASDU类型",
        ]
    )

    if protocol == "modbus":
        sample_yc = [
            ["PCS_DC_V", "直流母线电压", 1, "0x0000", 3, "0x41", 0.1, 0, 1000, 0],
            ["PCS_DC_I", "直流母线电流", 1, "0x0002", 3, "0x41", 0.1, 0, 500, -500],
            ["PCS_AC_P", "交流有功功率", 1, "0x0004", 3, "0x41", 0.1, 0, 1000, -1000],
            ["PCS_AC_Q", "交流无功功率", 1, "0x0006", 3, "0x41", 0.1, 0, 500, -500],
            ["BMS_SOC", "电池SOC", 2, "0x0000", 3, "0x41", 0.1, 0, 100, 0],
            ["BMS_TEMP", "电池温度", 2, "0x0002", 3, "0x41", 0.1, 0, 60, -20],
        ]
    elif protocol == "iec104":
        sample_yc = [
            ["PCS_DC_V", "直流母线电压", 1, "16385", 0, "", 0.1, 0, 1000, 0, "M_ME_NC_1"],
            ["PCS_DC_I", "直流母线电流", 1, "16386", 0, "", 0.1, 0, 500, -500, "M_ME_NC_1"],
            ["PCS_AC_P", "交流有功功率", 1, "16387", 0, "", 0.1, 0, 1000, -1000, "M_ME_NC_1"],
            ["BMS_SOC", "电池SOC", 1, "16388", 0, "", 0.1, 0, 100, 0, "M_ME_NC_1"],
        ]
    else:  # dlt645
        sample_yc = [
            ["METER_E_P", "正向有功电能", 1, "0x00010000", 3, "0x41", 0.01, 0, 999999, 0],
            ["METER_E_N", "反向有功电能", 1, "0x00020000", 3, "0x41", 0.01, 0, 999999, 0],
            ["METER_V_A", "A相电压", 1, "0x02010100", 3, "0x41", 0.1, 0, 300, 0],
            ["METER_I_A", "A相电流", 1, "0x02020100", 3, "0x41", 0.001, 0, 100, 0],
        ]

    for row in sample_yc:
        ws_yc.append(row)

    # ===== 遥信 sheet =====
    ws_yx = wb.create_sheet("遥信")
    ws_yx.append(
        ["测点编码", "测点名称", "从机地址", "寄存器地址", "功能码", "解析码", "位偏移", "是否反转", "104ASDU类型"]
    )

    if protocol == "modbus":
        sample_yx = [
            ["PCS_RUN", "PCS运行状态", 1, "0x0100", 1, "0x20", 0, 0],
            ["PCS_FAULT", "PCS故障", 1, "0x0100", 1, "0x20", 1, 0],
            ["PCS_CHARGE", "充电状态", 1, "0x0100", 1, "0x20", 2, 0],
            ["BMS_RUN", "BMS运行状态", 2, "0x0100", 1, "0x20", 0, 0],
            ["BMS_FAULT", "BMS故障", 2, "0x0100", 1, "0x20", 1, 0],
        ]
    elif protocol == "iec104":
        sample_yx = [
            ["PCS_RUN", "PCS运行状态", 1, "1", 0, "", 0, 0, "M_SP_NA_1"],
            ["PCS_FAULT", "PCS故障", 1, "2", 0, "", 0, 0, "M_SP_NA_1"],
            ["BMS_RUN", "BMS运行状态", 1, "3", 0, "", 0, 0, "M_SP_NA_1"],
        ]
    else:
        sample_yx = []

    for row in sample_yx:
        ws_yx.append(row)

    # ===== 遥控 sheet =====
    ws_yk = wb.create_sheet("遥控")
    ws_yk.append(
        [
            "测点编码",
            "测点名称",
            "从机地址",
            "寄存器地址",
            "功能码",
            "解析码",
            "位偏移",
            "命令类型",
            "关联遥信编码",
            "104ASDU类型",
        ]
    )

    if protocol == "modbus":
        sample_yk = [
            ["PCS_START", "PCS启动", 1, "0x0200", 5, "0x20", 0, 0, "PCS_RUN"],
            ["PCS_STOP", "PCS停止", 1, "0x0201", 5, "0x20", 0, 0, "PCS_RUN"],
            ["PCS_RESET", "PCS复位", 1, "0x0202", 5, "0x20", 0, 0, "PCS_FAULT"],
        ]
    elif protocol == "iec104":
        sample_yk = [
            ["PCS_START", "PCS启动", 1, "1", 0, "", 0, 0, "PCS_RUN", "C_SC_NA_1"],
            ["PCS_STOP", "PCS停止", 1, "2", 0, "", 0, 0, "PCS_RUN", "C_SC_NA_1"],
        ]
    else:
        sample_yk = []

    for row in sample_yk:
        ws_yk.append(row)

    # ===== 遥调 sheet =====
    ws_yt = wb.create_sheet("遥调")
    ws_yt.append(
        [
            "测点编码",
            "测点名称",
            "从机地址",
            "寄存器地址",
            "功能码",
            "解析码",
            "乘系数",
            "加系数",
            "上限值",
            "下限值",
            "关联遥测编码",
            "104ASDU类型",
        ]
    )

    if protocol == "modbus":
        sample_yt = [
            ["PCS_P_SET", "有功功率设定", 1, "0x0300", 6, "0x41", 0.1, 0, 1000, -1000, "PCS_AC_P"],
            ["PCS_Q_SET", "无功功率设定", 1, "0x0302", 6, "0x41", 0.1, 0, 500, -500, "PCS_AC_Q"],
            ["BMS_SOC_UP", "SOC上限设定", 2, "0x0300", 6, "0x41", 0.1, 0, 100, 0, "BMS_SOC"],
            ["BMS_SOC_LOW", "SOC下限设定", 2, "0x0302", 6, "0x41", 0.1, 0, 100, 0, "BMS_SOC"],
        ]
    elif protocol == "iec104":
        sample_yt = [
            ["PCS_P_SET", "有功功率设定", 1, "1", 0, "", 0.1, 0, 1000, -1000, "PCS_AC_P", "C_SE_NC_1"],
            ["PCS_Q_SET", "无功功率设定", 1, "2", 0, "", 0.1, 0, 500, -500, "", "C_SE_NC_1"],
        ]
    else:
        sample_yt = []

    for row in sample_yt:
        ws_yt.append(row)

    wb.save(file_path)
    wb.close()
    print(f"示例文件已创建: {file_path}")


if __name__ == "__main__":
    # 测试创建模板和示例文件
    create_excel_template("point_template.xlsx")
    create_sample_excel("point_sample_modbus.xlsx", "modbus")
    create_sample_excel("point_sample_iec104.xlsx", "iec104")
    create_sample_excel("point_sample_dlt645.xlsx", "dlt645")
