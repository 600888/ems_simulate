import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from src.tools.excel_point_importer import ExcelPointImporter


def _sheet(headers: list[str], values: list[object]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "遥测"
    sheet.append(headers)
    sheet.append(values)
    return workbook, sheet, tuple(values)


def test_dnp3_columns_are_parsed_into_validated_point_config() -> None:
    headers = [
        "测点编码",
        "DNP3点位类型",
        "静态变体",
        "事件变体",
        "CLASS",
        "死区",
        "初始品质",
        "产生事件",
        "事件携带时标",
    ]
    values = ["AI_1", "Analog Input", "V5", 7, "Class 3", 0.25, "0x01", 1, 0]
    workbook, sheet, row = _sheet(headers, values)

    raw = ExcelPointImporter._parse_dnp3_config(sheet, row, frame_type=0, excel_row=2)

    assert raw is not None
    config = json.loads(raw)
    assert config["static_variation"] == 5
    assert config["event_variation"] == 7
    assert config["event_class"] == 3
    assert config["deadband"] == 0.25
    assert config["initial_quality"] == 1
    assert config["event_enabled"] is True
    assert config["timestamp_enabled"] is False
    workbook.close()


@pytest.mark.parametrize("invalid_value", ["是", "否", True, False, "1", "0", 2, -1])
def test_dnp3_boolean_columns_strictly_require_numeric_zero_or_one(invalid_value: object) -> None:
    headers = ["测点编码", "DNP3点位类型", "产生事件", "事件携带时标"]
    workbook, sheet, row = _sheet(headers, ["AI_1", "Analog Input", invalid_value, 1])

    with pytest.raises(ValueError, match="产生事件必须填写数值 0 或 1"):
        ExcelPointImporter._parse_dnp3_config(sheet, row, frame_type=0, excel_row=2)

    workbook.close()


def test_dnp3_point_type_must_match_sheet_frame_type() -> None:
    workbook, sheet, row = _sheet(["测点编码", "DNP3点位类型"], ["AI_1", "Binary Input"])

    with pytest.raises(ValueError, match="Analog Input"):
        ExcelPointImporter._parse_dnp3_config(sheet, row, frame_type=0, excel_row=2)

    workbook.close()


def test_dnp3_point_type_is_not_imported_as_iec104_type() -> None:
    workbook, sheet, row = _sheet(["测点编码", "DNP3点位类型"], ["AI_1", "Analog Input"])

    assert ExcelPointImporter._iec104_type(sheet, row) is None

    workbook.close()


def test_shared_dnp3_sample_contains_no_iec104_columns_and_all_rows_parse() -> None:
    sample_path = Path(__file__).parents[2] / "data" / "point_csv" / "point_sample_dnp3.xlsx"
    workbook = load_workbook(sample_path, data_only=True)
    frame_types = {"遥测": 0, "遥信": 1, "遥控": 2, "遥调": 3}

    assert workbook.sheetnames == list(frame_types)
    for sheet_name, frame_type in frame_types.items():
        sheet = workbook[sheet_name]
        headers = ExcelPointImporter._header_indices(sheet)
        assert "104ASDU类型" not in headers
        assert "IEC104类型" not in headers
        assert {"DNP3点位类型", "静态变体", "事件变体", "CLASS"}.issubset(headers)
        for excel_row, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            assert row[headers["产生事件"]] == (1 if frame_type in (0, 1) else 0)
            assert row[headers["事件携带时标"]] == 1
            assert ExcelPointImporter._parse_dnp3_config(sheet, row, frame_type, excel_row)

    workbook.close()
